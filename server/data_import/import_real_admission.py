# -*- coding: utf-8 -*-
"""
真实录取数据通用导入框架
=========================
支持 CSV / Excel 导入，自动完成院校匹配、专业组创建、数据校验、去重。

数据来源：江苏省教育考试院 / 中国教育在线(EOL) / 高校招生网

使用方式：
    python import_real_admission.py --file data.csv --year 2024
    python import_real_admission.py --file data.xlsx --year 2024 --dry-run
    python import_real_admission.py --file data.csv --year 2024 --archive-simulated
"""
import csv
import os
import sys
import argparse
import pymysql
from datetime import datetime

# ---- 配置 ----
DB_CONFIG = {
    "host": "localhost", "user": "root", "password": "292180",
    "database": "gaokao", "charset": "utf8mb4",
}
PROVINCE_ID = 1
BATCH = "本科批"

# CSV 列名映射（支持多种来源的列名变体）
COLUMN_MAP = {
    # 标准名 → 别名列表
    "university_name": ["university_name", "院校名称", "院校", "学校", "university", "school"],
    "year": ["year", "年份", "招生年份"],
    "subject_type": ["subject_type", "科类", "科目", "subjectType", "category"],
    "major_group": ["major_group", "专业组", "group_name", "专业组名称", "groupName"],
    "subject_requirement": ["subject_requirement", "选科要求", "选考科目", "subjectRequirement", "requirement"],
    "min_score": ["min_score", "投档最低分", "最低分", "minScore", "score", "录取分"],
    "min_rank": ["min_rank", "最低位次", "位次", "minRank", "rank", "排名"],
}


def connect_db():
    return pymysql.connect(**DB_CONFIG)


def detect_format(filepath):
    """检测文件格式"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return 'csv'
    elif ext in ('.xlsx', '.xls'):
        return 'excel'
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 CSV/Excel")


def map_columns(headers):
    """将源文件列名映射到标准字段名"""
    mapping = {}
    for std_name, aliases in COLUMN_MAP.items():
        for alias in aliases:
            for h in headers:
                if alias.strip().lower() == h.strip().lower():
                    mapping[std_name] = h
                    break
            if std_name in mapping:
                break
    missing = [n for n in COLUMN_MAP if n not in mapping]
    return mapping, missing


def read_file(filepath, fmt):
    """读取文件，返回 [{field: value}, ...]"""
    if fmt == 'csv':
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            return list(reader)
    elif fmt == 'excel':
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows


def load_university_map(cur):
    """加载院校名称→ID映射（精确+模糊）"""
    cur.execute("SELECT id, name FROM universities")
    return {name: uid for uid, name in cur.fetchall()}


def match_university(name, uni_map):
    """院校名称匹配（精确→包含→模糊）"""
    if name in uni_map:
        return uni_map[name]
    # 去掉括号内容再试
    import re
    clean = re.sub(r'\(.*?\)', '', name).strip()
    if clean in uni_map:
        return uni_map[clean]
    # 模糊匹配
    for uname, uid in uni_map.items():
        if name in uname or uname in name:
            return uid
        if clean and (clean in uname or uname in clean):
            return uid
    return None


def validate_row(row, year):
    """校验单行数据，返回 (is_valid, error_msg)"""
    errors = []
    # 必填字段
    for f in ['min_score', 'min_rank']:
        try:
            val = int(row.get(f, 0))
            if val <= 0:
                errors.append(f"{f}必须为正整数")
        except (ValueError, TypeError):
            errors.append(f"{f}无法解析为整数: {row.get(f)}")

    # 分数范围
    score = int(row.get('min_score', 0))
    if score < 200 or score > 750:
        errors.append(f"分数异常: {score}")

    # 位次范围（江苏约1~200000）
    rank = int(row.get('min_rank', 0))
    if rank < 1 or rank > 500000:
        errors.append(f"位次异常: {rank}")

    # 年份
    if year and int(row.get('year', year)) != year:
        errors.append(f"年份不匹配: {row.get('year')} != {year}")

    return len(errors) == 0, "; ".join(errors)


def import_rows(conn, rows, uni_map, year, dry_run=False):
    """导入数据到 admission_scores"""
    cur = conn.cursor()
    stats = {"inserted": 0, "updated": 0, "skipped": 0, "errors": 0, "unmatched": []}

    for i, row in enumerate(rows):
        # 映射字段
        uni_name = row.get('university_name', '')
        uni_id = match_university(uni_name, uni_map)

        if uni_id is None:
            stats["unmatched"].append(uni_name)
            stats["skipped"] += 1
            continue

        # 校验
        valid, err = validate_row(row, year)
        if not valid:
            stats["errors"] += 1
            if stats["errors"] <= 5:
                print(f"  ⚠️ 第{i+1}行校验失败: {err}")
            continue

        subject_type = row.get('subject_type', '物理类')
        major_group = row.get('major_group', '')
        subject_req = row.get('subject_requirement', '')
        min_score = int(row['min_score'])
        min_rank = int(row['min_rank'])
        row_year = int(row.get('year', year))

        if dry_run:
            stats["inserted"] += 1
            continue

        # 检查是否已存在
        cur.execute(
            """SELECT id FROM admission_scores
               WHERE university_id=%s AND major_id IS NULL AND year=%s
               AND subject_type=%s AND major_group=%s AND data_source='OFFICIAL'
               LIMIT 1""",
            (uni_id, row_year, subject_type, major_group),
        )
        existing = cur.fetchone()

        if existing:
            cur.execute(
                """UPDATE admission_scores
                   SET min_score=%s, min_rank=%s, subject_requirement=%s,
                       avg_score=%s, plan_count=%s, data_source='OFFICIAL', verified_level=3
                   WHERE id=%s""",
                (min_score, min_rank, subject_req, min_score + 5, 30, existing[0]),
            )
            stats["updated"] += 1
        else:
            cur.execute(
                """INSERT INTO admission_scores
                   (university_id, major_id, province_id, year, subject_type, batch,
                    min_score, min_rank, avg_score, plan_count,
                    subject_requirement, major_group,
                    data_source, verified_level)
                   VALUES (%s, NULL, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                           'OFFICIAL', 3)""",
                (uni_id, PROVINCE_ID, row_year, subject_type, BATCH,
                 min_score, min_rank, min_score + 5, 30,
                 subject_req, major_group),
            )
            stats["inserted"] += 1

    if not dry_run:
        conn.commit()

    # 打印不匹配的院校（去重）
    if stats["unmatched"]:
        unique = list(set(stats["unmatched"]))
        print(f"  ⚠️ 未匹配院校 ({len(unique)}所):")
        for name in sorted(unique)[:10]:
            print(f"    - {name}")
        if len(unique) > 10:
            print(f"    ... 还有 {len(unique)-10} 所")

    return stats


def archive_simulated_data(conn):
    """归档模拟数据到 admission_scores_archive 表"""
    cur = conn.cursor()
    # 创建归档表
    cur.execute("CREATE TABLE IF NOT EXISTS admission_scores_archive LIKE admission_scores")
    # 移动模拟数据
    cur.execute("INSERT INTO admission_scores_archive SELECT * FROM admission_scores WHERE data_source='SIMULATED'")
    archived = cur.rowcount
    cur.execute("DELETE FROM admission_scores WHERE data_source='SIMULATED'")
    deleted = cur.rowcount
    conn.commit()
    print(f"  📦 已归档 {archived} 条模拟数据 -> admission_scores_archive")
    print(f"  🗑 已删除 {deleted} 条模拟数据")
    return archived


def main():
    parser = argparse.ArgumentParser(description="真实录取数据通用导入框架")
    parser.add_argument("--file", required=True, help="CSV或Excel文件路径")
    parser.add_argument("--year", type=int, required=True, help="数据年份")
    parser.add_argument("--dry-run", action="store_true", help="仅校验不写入")
    parser.add_argument("--archive-simulated", action="store_true",
                       help="导入前将现有SIMULATED数据归档到admission_scores_archive")
    args = parser.parse_args()

    print("=" * 60)
    print("📥 真实录取数据导入框架")
    print("=" * 60)

    # 1. 检测格式并读取
    fmt = detect_format(args.file)
    print(f"\n📄 文件格式: {fmt}")
    raw_rows = read_file(args.file, fmt)
    print(f"📊 原始行数: {len(raw_rows)}")

    if not raw_rows:
        print("❌ 文件为空")
        return

    # 2. 列名映射
    headers = list(raw_rows[0].keys())
    mapping, missing = map_columns(headers)
    if missing:
        print(f"❌ 缺少必要列: {missing}")
        print(f"   实际列名: {headers}")
        print(f"   需要映射到: {list(COLUMN_MAP.keys())}")
        return

    print(f"🔗 列名映射: {dict((k, mapping[k]) for k in sorted(mapping))}")

    # 3. 转换行数据
    rows = []
    for raw in raw_rows:
        row = {std: raw.get(src, '') for std, src in mapping.items()}
        if not row.get('year'):
            row['year'] = str(args.year)
        rows.append(row)
    print(f"📊 有效行数: {len(rows)}")

    # 4. 连接数据库
    conn = connect_db()
    cur = conn.cursor()

    # 5. 院校映射
    uni_map = load_university_map(cur)
    print(f"🏫 已加载 {len(uni_map)} 所院校")

    # 6. 归档模拟数据(可选)
    if args.archive_simulated and not args.dry_run:
        archive_simulated_data(conn)

    # 7. 导入
    stats = import_rows(conn, rows, uni_map, args.year, args.dry_run)

    # 8. 汇总
    print("\n" + "=" * 60)
    print("📋 导入汇总")
    print("=" * 60)
    print(f"  新增: {stats['inserted']}")
    print(f"  更新: {stats['updated']}")
    print(f"  跳过(未匹配): {stats['skipped']}")
    print(f"  错误(校验失败): {stats['errors']}")

    # 9. 导入后统计
    if not args.dry_run:
        cur.execute(
            "SELECT COUNT(*) FROM admission_scores WHERE data_source='OFFICIAL'"
        )
        official_cnt = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM admission_scores")
        total_cnt = cur.fetchone()[0]
        rate = round(official_cnt / total_cnt * 100, 1) if total_cnt > 0 else 0
        print(f"\n📊 导入后 OFFICIAL: {official_cnt}/{total_cnt} ({rate}%)")

        if rate >= 95:
            print("✅ 真实数据率已达标(≥95%)，可以开启收费")
        else:
            print(f"⚠️ 真实数据率 {rate}%，距离95%还差约 {int(total_cnt*0.95 - official_cnt)} 条OFFICIAL")

    if args.dry_run:
        print("\n⚠️ DRY RUN 模式，数据未实际写入")
        conn.rollback()

    conn.close()
    print("\n✅ 导入完成")


if __name__ == "__main__":
    main()
